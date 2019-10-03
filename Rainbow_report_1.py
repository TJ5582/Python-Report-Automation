
#### OUTPUT WRIITTEN DIRECTLY INTO RAINBOW REPORT EXCEL FILE

import pandas as pd
import numpy as np
from datetime import date, timedelta, datetime

# libraries to be imported 
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders
from datetime import date, timedelta

import openpyxl

#### Disabling SettingwithCopyWarning 
pd.set_option('mode.chained_assignment', None)

data_frame = pd.read_excel('mod_Daily App Flyer data with City and Way2Online.xlsx', sheet_name='FINAL_DAILY')


start_date = date.today().strftime("01-%b-19")
start_date = datetime.strptime(start_date, '%d-%b-%y').date()
start_date_str = start_date.strftime('%d-%b-%y')
print (start_date)
#print (start_date_str)


today = date.today()
report_date = today.strftime("%d-%b-%y")
print("report_date=", report_date)
yest_date = date.today() - timedelta(days=1)
yesterday_date = yest_date.strftime("%d-%b-%y")
print("yesterday date =", yesterday_date)



#yesterday_date = '01-Aug-19'
day = today.strftime("%a")
print (day)

month = today.strftime("%a")
#print (month)


#day = 'Mon'

if (day != 'Mon'):
	df_mod = data_frame[(data_frame['Loan_Type'] != 'BL') & (data_frame['Created date'] == yesterday_date)]
	num_of_days = 1
	dd_date = (date.today() - timedelta(days=1)).strftime("%d-%b-%y")
	yesterday_day = (date.today() - timedelta(days=1)).strftime("%a")


else:
	last_friday_date = (date.today() - timedelta(days=3)).strftime("%d-%b-%y")
	#print (last_friday_date)
	data_frame = data_frame[(data_frame['Loan_Type'] != 'BL') & (data_frame['Created date'] >= last_friday_date)]
	fri_date = date.today() - timedelta(days=3)
	num_of_days = (today - fri_date)
	#print ("num_of_days = ",num_of_days.days)
	num_of_days = num_of_days.days


#num_of_days = 2

counter = num_of_days
index = 0
html_str =''
df_html=pd.DataFrame()
start_row_idx_1 = 0
start_row_idx_2 = 0
start_row_idx_3 = 0
start_row_idx_4 = 0



while (counter > 0):

	dd_date_date = date.today() - timedelta(days=counter)
	dd_date = (date.today() - timedelta(days=counter)).strftime("%d-%b-%y")
	print (dd_date)
	ndays = dd_date_date - start_date
	ndays = ndays.days
	print (ndays)

	start_row_idx_1 = 6 + ndays
	start_row_idx_2 = 41 + ndays
	start_row_idx_3 = 76 + ndays
	start_row_idx_4 = 112 + ndays

	print (start_row_idx_1,start_row_idx_2,start_row_idx_3,start_row_idx_4)

	
	df_mod = data_frame[(data_frame['Loan_Type'] != 'BL') & (data_frame['Created date'] == dd_date)]
	#print df_mod

	yesterday_day = (date.today() - timedelta(days=counter)).strftime("%a")



	df_mod["Media_cost"].fillna("Organic", inplace = True)
	df_mod["Partner"].fillna("", inplace = True)

	
	print("########## Application sum ORGANIC ######################")


	df1 = df_mod.loc[df_mod['Partner'] != 'inmobiagen',:]

	app_sum_org_2 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	app_sum_org_2.reset_index(inplace=True)

	app_blanks = list(app_sum_org_2.loc[app_sum_org_2["Media_cost"] == 'Organic','Total'])
	
	if(app_blanks == []):
		app_blanks.append(0)

	print (app_blanks[0])


	print("########## Disbursal Count ORGANIC ######################")

	disbursal_sum_org_2 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	disbursal_sum_org_2.reset_index(inplace=True)

	disbursal_blanks = list(disbursal_sum_org_2.loc[disbursal_sum_org_2["Media_cost"] == 'Organic','Total'])

	if(disbursal_blanks == []):
		disbursal_blanks.append(0)

	print (disbursal_blanks[0])

	print("########## Disbursal AMOUNT ORGANIC ######################")

	disbursal_amount_org_2 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	disbursal_amount_org_2.reset_index(inplace=True)

	disbursal_amt_blanks = list(disbursal_amount_org_2.loc[disbursal_amount_org_2["Media_cost"] == 'Organic','Total'])
	
	if(disbursal_amt_blanks == []):
		disbursal_amt_blanks.append(0)

	print (disbursal_amt_blanks[0])

	print("########## GOOGLE UAC = Madison ######################")

	df1 = df_mod.loc[(df_mod['Partner'] == 'adapptmobi') | (df_mod['Partner'] == 'inmobiagen') | (df_mod['Partner'] == 'madison') | (df_mod['Partner'] == 'vserv'),:]

	if (df1.empty == False):
		app_sum_org_1 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		m_google_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','count_dis_Allstage']
		m_google_app_count_df.columns = ['Count']
		m_google_app_count_df.reset_index(inplace=True)
		m_google_app_count = m_google_app_count_df['Count'][0]
		print (m_google_app_count)

		m_google_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','disb_cust']
		m_google_disbursal_count_df.columns = ['Count']
		m_google_disbursal_count_df.reset_index(inplace=True)
		m_google_disbursal_count = m_google_disbursal_count_df['Count'][0]
		print (m_google_disbursal_count)
		
		m_google_disbursal_amount_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','disb_amount']
		m_google_disbursal_amount_df.columns = ['Count']
		m_google_disbursal_amount_df.reset_index(inplace=True)
		m_google_disbursal_amount = m_google_disbursal_amount_df['Count'][0]
		print(m_google_disbursal_amount)
	
		print("########## FACEBOOK  = Madison ######################")
		m_fb_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Facebook','count_dis_Allstage']
		m_fb_app_count_df.columns = ['Count']
		m_fb_app_count_df.reset_index(inplace=True)
		m_fb_app_count = m_fb_app_count_df['Count'][0]
		print (m_fb_app_count)

		m_fb_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Facebook','disb_cust']
		m_fb_disbursal_count_df.columns = ['Count']
		m_fb_disbursal_count_df.reset_index(inplace=True)
		m_fb_disbursal_count = m_fb_disbursal_count_df['Count'][0]
		print (m_fb_disbursal_count)

		m_fb_disbursal_amount_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Facebook','disb_amount']
		m_fb_disbursal_amount_df.columns = ['Count']
		m_fb_disbursal_amount_df.reset_index(inplace=True)
		m_fb_disbursal_amount = m_fb_disbursal_amount_df['Count'][0]
		print (m_fb_disbursal_amount)

		print("########## FACEBOOK  ADS = Madison ######################")
		m_fbads_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting','count_dis_Allstage']
		m_fbads_app_count_df.columns = ['Count']
		m_fbads_app_count_df.reset_index(inplace=True)
		m_fbads_app_count = m_fbads_app_count_df['Count'][0]
		print (m_fbads_app_count)

		m_fbads_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting','disb_cust']
		m_fbads_disbursal_count_df.columns = ['Count']
		m_fbads_disbursal_count_df.reset_index(inplace=True)
		m_fbads_disbursal_count = m_fbads_disbursal_count_df['Count'][0]
		print (m_fbads_disbursal_count)

		m_fbads_disbursal_amount_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting','disb_amount']
		m_fbads_disbursal_amount_df.columns = ['Count']
		m_fbads_disbursal_amount_df.reset_index(inplace=True)
		m_fbads_disbursal_amount = m_fbads_disbursal_amount_df['Count'][0]
		print (m_fbads_disbursal_amount)

	else:
		m_google_app_count = 0
		m_google_disbursal_count = 0
		m_google_disbursal_amount = 0

		m_fb_app_count = 0
		m_fb_disbursal_count = 0
		m_fb_disbursal_amount = 0

		m_fbads_app_count = 0
		m_fbads_disbursal_count = 0
		m_fbads_disbursal_amount = 0


	print("########## GOOGLE UAC = Sokrati ######################")

	df2 = df_mod.loc[(df_mod['Partner'] == '') | (df_mod['Partner'] == 'sokrati'),:]

	if (df2.empty == False):
		app_sum_org_1 = pd.pivot_table(df2, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		sk_google_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','count_dis_Allstage']
		sk_google_app_count_df.columns = ['Count']
		sk_google_app_count_df.reset_index(inplace=True)
		sk_google_app_count = sk_google_app_count_df['Count'][0]
		print (sk_google_app_count)

		sk_google_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','disb_cust']
		sk_google_disbursal_count_df.columns = ['Count']
		sk_google_disbursal_count_df.reset_index(inplace=True)
		sk_google_disbursal_count = sk_google_disbursal_count_df['Count'][0]
		print (sk_google_disbursal_count)

		sk_google_disbursal_amount_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Google UAC','disb_amount']
		sk_google_disbursal_amount_df.columns = ['Count']
		sk_google_disbursal_amount_df.reset_index(inplace=True)
		sk_google_disbursal_amount = sk_google_disbursal_amount_df['Count'][0]
		print (sk_google_disbursal_amount)
		
	else:
		sk_google_app_count = 0
		sk_google_disbursal_count = 0
		sk_google_disbursal_amount = 0

	print("########## FACEBOOK = Sokrati ######################")

	df3 = df_mod.loc[df_mod['Partner'] == 'sokrati',:]

	if (df3.empty == False):
		
		app_sum_org_1 = pd.pivot_table(df3, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		sk_fb_app_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'count_dis_Allstage']
		sk_fb_app_count_df.columns = ['Count']
		sk_fb_app_count_df.reset_index(inplace=True)
		print (sk_fb_app_count_df)
		sk_fb_app_count = sk_fb_app_count_df['Count'][0]
		print (sk_fb_app_count)

		sk_fb_disbursal_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'disb_cust']
		sk_fb_disbursal_count_df.columns = ['Count']
		sk_fb_disbursal_count_df.reset_index(inplace=True)
		print (sk_fb_disbursal_count_df)
		sk_fb_disbursal_count = sk_fb_disbursal_count_df['Count'][0]
		print (sk_fb_disbursal_count)

		sk_fb_disbursal_amount_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'disb_amount']
		sk_fb_disbursal_amount_df.columns = ['Count']
		sk_fb_disbursal_amount_df.reset_index(inplace=True)
		print (sk_fb_disbursal_amount_df)
		sk_fb_disbursal_amount = sk_fb_disbursal_amount_df['Count'][0]
		print (sk_fb_disbursal_amount)


		print("########## FACEBOOK ADs = Sokrati ######################")

		if (sk_fb_app_count_df.last_valid_index() == 1):
			sk_fbads_app_count = sk_fb_app_count_df['Count'][1]
		else:
			sk_fbads_app_count = 0

		if (sk_fb_disbursal_count_df.last_valid_index() == 1):
			sk_fbads_disbursal_count = sk_fb_disbursal_count_df['Count'][1]
		else:
			sk_fbads_disbursal_count = 0

		if (sk_fb_disbursal_amount_df.last_valid_index() == 1):
			sk_fbads_disbursal_amount = sk_fb_disbursal_amount_df['Count'][1]
		else:
			sk_fbads_disbursal_amount = 0
		
		print (sk_fbads_app_count)
		print (sk_fbads_disbursal_count)
		print (sk_fbads_disbursal_amount)

		sk_fb_app_tot_count = sk_fb_app_count + sk_fbads_app_count
		sk_fb_disbursal_tot_count = sk_fb_disbursal_count + sk_fbads_disbursal_count
		sk_fb_disbursal_tot_amount = sk_fb_disbursal_amount + sk_fbads_disbursal_amount

		print("########## TOTAL FACEBOOK = Sokrati ######################")
		print (sk_fb_app_tot_count)
		print (sk_fb_disbursal_tot_count)
		print (sk_fb_disbursal_tot_amount)

	print("########## VALUE LEAF  ######################")

	df3 = df_mod.loc[df_mod['Partner'] == '',:]

	if (df3.empty == False):

		app_sum_org_1 = pd.pivot_table(df3, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		VL_app_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'count_dis_Allstage']

		if (VL_app_count_df.empty == False):
			
			VL_app_count_df.columns = ['Count']
			VL_app_count_df.reset_index(inplace=True)
			#print (VL_app_count_df)

			if (VL_app_count_df.last_valid_index() == 1):

				VL_fb_app_count = VL_app_count_df['Count'][0]
				VL_fbads_app_count = VL_app_count_df['Count'][1]

			else:
				VL_fb_app_count = VL_app_count_df['Count'][0]
				VL_fbads_app_count = 0

			VL_tot_app_count = VL_fb_app_count + VL_fbads_app_count

		else:
			VL_tot_app_count = 0

	
		

		VL_disbursal_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'disb_cust']
	
		if (VL_disbursal_count_df.empty == False):
			
			VL_disbursal_count_df.columns = ['Count']
			VL_disbursal_count_df.reset_index(inplace=True)

			#print (VL_disbursal_count_df)

			if (VL_disbursal_count_df.last_valid_index() == 1):
				
				VL_fb_tot_disbursal_count = VL_disbursal_count_df['Count'][0]
				VL_fbads_tot_disbursal_count = VL_disbursal_count_df['Count'][1]

			else:
				VL_fb_tot_disbursal_count = VL_disbursal_count_df['Count'][0]
				VL_fbads_tot_disbursal_count = 0

			VL_tot_disbursal_count = VL_fb_tot_disbursal_count + VL_fbads_tot_disbursal_count

		else:
			VL_tot_disbursal_count = 0


		VL_disbursal_amount_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'Facebook') | (app_sum_org_1["Media_cost"] == 'FacebookAds_Retargetting'),'disb_amount']
	
		if (VL_disbursal_amount_df.empty == False):
			
			VL_disbursal_amount_df.columns = ['Count']
			VL_disbursal_amount_df.reset_index(inplace=True)

			if (VL_disbursal_amount_df.last_valid_index() == 1):

				VL_fb_tot_disbursal_amount = VL_disbursal_amount_df['Count'][0]
				VL_fbads_tot_disbursal_amount = VL_disbursal_amount_df['Count'][1]
			
			else:
				VL_fb_tot_disbursal_amount = VL_disbursal_amount_df['Count'][0]
				VL_fbads_tot_disbursal_amount = 0


			VL_tot_disbursal_amount = VL_fb_tot_disbursal_amount + VL_fbads_tot_disbursal_amount

		else:
			VL_tot_disbursal_amount = 0

		print(VL_tot_app_count,VL_tot_disbursal_count,VL_tot_disbursal_amount)


	print("########## inmobi APP COUNT ######################")
	
	#### Using df_mod
	
	app_sum_org_1 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
	app_sum_org_1.reset_index(inplace=True)
	#app_sum_org_1.to_csv("app_sum_org_1.csv")

	m_inmobi_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'InMobi','count_dis_Allstage']

	if (m_inmobi_app_count_df.empty == False):

		m_inmobi_app_count_df.columns = ['Count']
		m_inmobi_app_count_df.reset_index(inplace=True)
		m_inmobi_app_count = m_inmobi_app_count_df['Count'][0]
	

	else:
		m_inmobi_app_count = 0

	m_inmobi_retarg_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'inmobi_int_Retargetting','count_dis_Allstage']

	if (m_inmobi_retarg_app_count_df.empty == False):

		m_inmobi_retarg_app_count_df.columns = ['Count']
		m_inmobi_retarg_app_count_df.reset_index(inplace=True)
		m_inmobi_retarg_app_count = m_inmobi_retarg_app_count_df['Count'][0]
	

	else:
		m_inmobi_retarg_app_count = 0

	m_inmobi_tot_app_count = m_inmobi_app_count + m_inmobi_retarg_app_count
	print(m_inmobi_app_count,m_inmobi_retarg_app_count,m_inmobi_tot_app_count)

	print("########## inmobi Disbursal COUNT ######################")


	m_inmobi_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'InMobi','disb_cust']

	if (m_inmobi_disbursal_count_df.empty == False):

		m_inmobi_disbursal_count_df.columns = ['Count']
		m_inmobi_disbursal_count_df.reset_index(inplace=True)
		m_inmobi_disbursal_count = m_inmobi_disbursal_count_df['Count'][0]
	

	else:
		m_inmobi_disbursal_count = 0

	m_inmobi_retarg_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'inmobi_int_Retargetting','disb_cust']

	if (m_inmobi_retarg_disbursal_count_df.empty == False):

		m_inmobi_retarg_disbursal_count_df.columns = ['Count']
		m_inmobi_retarg_disbursal_count_df.reset_index(inplace=True)
		m_inmobi_retarg_disbursal_count = m_inmobi_retarg_disbursal_count_df['Count'][0]
	

	else:
		m_inmobi_retarg_disbursal_count = 0

	m_inmobi_tot_disbursal_count = m_inmobi_disbursal_count + m_inmobi_retarg_disbursal_count
	print(m_inmobi_disbursal_count,m_inmobi_retarg_disbursal_count,m_inmobi_tot_disbursal_count)


	print("########## inmobi Disbursal AMOUNT ######################")

	m_inmobi_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'InMobi','disb_amount']

	if (m_inmobi_disbursal_amt_count_df.empty == False):

		m_inmobi_disbursal_amt_count_df.columns = ['Count']
		m_inmobi_disbursal_amt_count_df.reset_index(inplace=True)
		m_inmobi_disbursal_amt_count = m_inmobi_disbursal_amt_count_df['Count'][0]
	

	else:
		m_inmobi_disbursal_amt_count = 0

	m_inmobi_retarg_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'inmobi_int_Retargetting','disb_amount']

	if (m_inmobi_retarg_disbursal_amt_count_df.empty == False):

		m_inmobi_retarg_disbursal_amt_count_df.columns = ['Count']
		m_inmobi_retarg_disbursal_amt_count_df.reset_index(inplace=True)
		m_inmobi_retarg_disbursal_amt_count = m_inmobi_retarg_disbursal_amt_count_df['Count'][0]
	

	else:
		m_inmobi_retarg_disbursal_amt_count = 0

	m_inmobi_tot_disbursal_amt_count = m_inmobi_disbursal_amt_count + m_inmobi_retarg_disbursal_amt_count
	print(m_inmobi_disbursal_amt_count,m_inmobi_retarg_disbursal_amt_count,m_inmobi_tot_disbursal_amt_count)

	


	print("########## pocket APP COUNT ######################")
	
	#### Using df_mod
	m_pocket_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Pocket','count_dis_Allstage']

	if (m_pocket_app_count_df.empty == False):

		m_pocket_app_count_df.columns = ['Count']
		m_pocket_app_count_df.reset_index(inplace=True)
		m_pocket_app_count = m_pocket_app_count_df['Count'][0]
	

	else:
		m_pocket_app_count = 0

	print(m_pocket_app_count)

	print("########## pocket Disbursal COUNT ######################")


	m_pocket_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Pocket','disb_cust']

	if (m_pocket_disbursal_count_df.empty == False):

		m_pocket_disbursal_count_df.columns = ['Count']
		m_pocket_disbursal_count_df.reset_index(inplace=True)
		m_pocket_disbursal_count = m_pocket_disbursal_count_df['Count'][0]
	

	else:
		m_pocket_disbursal_count = 0

	print(m_pocket_disbursal_count)


	print("########## pocket Disbursal AMOUNT ######################")

	m_pocket_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Pocket','disb_amount']

	if (m_pocket_disbursal_amt_count_df.empty == False):

		m_pocket_disbursal_amt_count_df.columns = ['Count']
		m_pocket_disbursal_amt_count_df.reset_index(inplace=True)
		m_pocket_disbursal_amt_count = m_pocket_disbursal_amt_count_df['Count'][0]
	

	else:
		m_pocket_disbursal_amt_count = 0


	print(m_pocket_disbursal_amt_count)


	print("########## LeadBolt APP COUNT ######################")
	
	#### Using df_mod
	m_LeadBolt_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'LeadBolt','count_dis_Allstage']

	if (m_LeadBolt_app_count_df.empty == False):

		m_LeadBolt_app_count_df.columns = ['Count']
		m_LeadBolt_app_count_df.reset_index(inplace=True)
		m_LeadBolt_app_count = m_LeadBolt_app_count_df['Count'][0]
	

	else:
		m_LeadBolt_app_count = 0

	print(m_LeadBolt_app_count)

	print("########## LeadBolt Disbursal COUNT ######################")


	m_LeadBolt_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'LeadBolt','disb_cust']

	if (m_LeadBolt_disbursal_count_df.empty == False):

		m_LeadBolt_disbursal_count_df.columns = ['Count']
		m_LeadBolt_disbursal_count_df.reset_index(inplace=True)
		m_LeadBolt_disbursal_count = m_LeadBolt_disbursal_count_df['Count'][0]
	

	else:
		m_LeadBolt_disbursal_count = 0

	print(m_LeadBolt_disbursal_count)


	print("########## LeadBolt Disbursal AMOUNT ######################")

	m_LeadBolt_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'LeadBolt','disb_amount']

	if (m_LeadBolt_disbursal_amt_count_df.empty == False):

		m_LeadBolt_disbursal_amt_count_df.columns = ['Count']
		m_LeadBolt_disbursal_amt_count_df.reset_index(inplace=True)
		m_LeadBolt_disbursal_amt_count = m_LeadBolt_disbursal_amt_count_df['Count'][0]
	

	else:
		m_LeadBolt_disbursal_amt_count = 0


	print(m_LeadBolt_disbursal_amt_count)

	print("########## Madison = Affiliates (Others) APP COUNT ######################")

	m_Aff_app_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'adapptmobi_int') | (app_sum_org_1["Media_cost"] == 'Adcanpus') | (app_sum_org_1["Media_cost"] == 'adcountymedia_int') | (app_sum_org_1["Media_cost"] == 'Adsplay') | (app_sum_org_1["Media_cost"] == 'advolt_int') | (app_sum_org_1["Media_cost"] == 'Affle') | (app_sum_org_1["Media_cost"] == 'appfloodaff_int') | (app_sum_org_1["Media_cost"] == 'appmontizemedia_int') | (app_sum_org_1["Media_cost"] == 'capslockdigitalsolutions') | (app_sum_org_1["Media_cost"] == 'claymotion_int') | (app_sum_org_1["Media_cost"] == 'Digital') | (app_sum_org_1["Media_cost"] == 'glispacpa_int') | (app_sum_org_1["Media_cost"] == 'iavatarzaffise_int') | (app_sum_org_1["Media_cost"] == 'Icubewires') | (app_sum_org_1["Media_cost"] == 'Intellectads') | (app_sum_org_1["Media_cost"] == 'mobidiscover_int') | (app_sum_org_1["Media_cost"] == 'mobireckon_int') | (app_sum_org_1["Media_cost"] == 'mobisummer_int') | (app_sum_org_1["Media_cost"] == 'mobpower2_int') | (app_sum_org_1["Media_cost"] == 'mobvista_int') | (app_sum_org_1["Media_cost"] == 'mobwonder_int') | (app_sum_org_1["Media_cost"] == 'omobiads_int') | (app_sum_org_1["Media_cost"] == 'pointific_int') | (app_sum_org_1["Media_cost"] == 'revx_int') | (app_sum_org_1["Media_cost"] == 'Sense') | (app_sum_org_1["Media_cost"] == 'svgmedia_int') | (app_sum_org_1["Media_cost"] == 'themobilyarabia_int') | (app_sum_org_1["Media_cost"] == 'Twitter') | (app_sum_org_1["Media_cost"] == 'tyroo_int') | (app_sum_org_1["Media_cost"] == 'uchuichuan_int') | (app_sum_org_1["Media_cost"] == 'vcommission_int') | (app_sum_org_1["Media_cost"] == 'vertozaff_int') | (app_sum_org_1["Media_cost"] == 'xaprio_int') | (app_sum_org_1["Media_cost"] == 'xyads_int') | (app_sum_org_1["Media_cost"] == 'yahoogemini_int'),'count_dis_Allstage']
	m_Aff_app_count_df.reset_index(inplace=True)
	print (m_Aff_app_count_df)

	m_Aff_app_total_count_df = 0
	
	i = 0

	while (i <= m_Aff_app_count_df.last_valid_index()):
		m_Aff_app_total_count_df = m_Aff_app_total_count_df + m_Aff_app_count_df.iloc[i,1]
		i = i + 1

	print (m_Aff_app_total_count_df)



	print("########## Madison = Affiliates (Others) Disbursal COUNT ######################")

	m_Aff_disbursal_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'adapptmobi_int') | (app_sum_org_1["Media_cost"] == 'Adcanpus') | (app_sum_org_1["Media_cost"] == 'adcountymedia_int') | (app_sum_org_1["Media_cost"] == 'Adsplay') | (app_sum_org_1["Media_cost"] == 'advolt_int') | (app_sum_org_1["Media_cost"] == 'Affle') | (app_sum_org_1["Media_cost"] == 'appfloodaff_int') | (app_sum_org_1["Media_cost"] == 'appmontizemedia_int') | (app_sum_org_1["Media_cost"] == 'capslockdigitalsolutions') | (app_sum_org_1["Media_cost"] == 'claymotion_int') | (app_sum_org_1["Media_cost"] == 'Digital') | (app_sum_org_1["Media_cost"] == 'glispacpa_int') | (app_sum_org_1["Media_cost"] == 'iavatarzaffise_int') | (app_sum_org_1["Media_cost"] == 'Icubewires') | (app_sum_org_1["Media_cost"] == 'Intellectads') | (app_sum_org_1["Media_cost"] == 'mobidiscover_int') | (app_sum_org_1["Media_cost"] == 'mobireckon_int') | (app_sum_org_1["Media_cost"] == 'mobisummer_int') | (app_sum_org_1["Media_cost"] == 'mobpower2_int') | (app_sum_org_1["Media_cost"] == 'mobvista_int') | (app_sum_org_1["Media_cost"] == 'mobwonder_int') | (app_sum_org_1["Media_cost"] == 'omobiads_int') | (app_sum_org_1["Media_cost"] == 'pointific_int') | (app_sum_org_1["Media_cost"] == 'revx_int') | (app_sum_org_1["Media_cost"] == 'Sense') | (app_sum_org_1["Media_cost"] == 'svgmedia_int') | (app_sum_org_1["Media_cost"] == 'themobilyarabia_int') | (app_sum_org_1["Media_cost"] == 'Twitter') | (app_sum_org_1["Media_cost"] == 'tyroo_int') | (app_sum_org_1["Media_cost"] == 'uchuichuan_int') | (app_sum_org_1["Media_cost"] == 'vcommission_int') | (app_sum_org_1["Media_cost"] == 'vertozaff_int') | (app_sum_org_1["Media_cost"] == 'xaprio_int') | (app_sum_org_1["Media_cost"] == 'xyads_int') | (app_sum_org_1["Media_cost"] == 'yahoogemini_int'),'disb_cust']
	m_Aff_disbursal_count_df.reset_index(inplace=True)
	#print (m_Aff_disbursal_count_df)

	m_Aff_disbursal_total_count_df = 0
	
	i = 0

	while (i <= m_Aff_disbursal_count_df.last_valid_index()):
		m_Aff_disbursal_total_count_df = m_Aff_disbursal_total_count_df + m_Aff_disbursal_count_df.iloc[i,1]
		i = i + 1

	print (m_Aff_disbursal_total_count_df)

	print("########## Madison = Affiliates (Others) Disbursal AMOUNT ######################")

	m_Aff_disbursal_amount_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'adapptmobi_int') | (app_sum_org_1["Media_cost"] == 'Adcanpus') | (app_sum_org_1["Media_cost"] == 'adcountymedia_int') | (app_sum_org_1["Media_cost"] == 'Adsplay') | (app_sum_org_1["Media_cost"] == 'advolt_int') | (app_sum_org_1["Media_cost"] == 'Affle') | (app_sum_org_1["Media_cost"] == 'appfloodaff_int') | (app_sum_org_1["Media_cost"] == 'appmontizemedia_int') | (app_sum_org_1["Media_cost"] == 'capslockdigitalsolutions') | (app_sum_org_1["Media_cost"] == 'claymotion_int') | (app_sum_org_1["Media_cost"] == 'Digital') | (app_sum_org_1["Media_cost"] == 'glispacpa_int') | (app_sum_org_1["Media_cost"] == 'iavatarzaffise_int') | (app_sum_org_1["Media_cost"] == 'Icubewires') | (app_sum_org_1["Media_cost"] == 'Intellectads') | (app_sum_org_1["Media_cost"] == 'mobidiscover_int') | (app_sum_org_1["Media_cost"] == 'mobireckon_int') | (app_sum_org_1["Media_cost"] == 'mobisummer_int') | (app_sum_org_1["Media_cost"] == 'mobpower2_int') | (app_sum_org_1["Media_cost"] == 'mobvista_int') | (app_sum_org_1["Media_cost"] == 'mobwonder_int') | (app_sum_org_1["Media_cost"] == 'omobiads_int') | (app_sum_org_1["Media_cost"] == 'pointific_int') | (app_sum_org_1["Media_cost"] == 'revx_int') | (app_sum_org_1["Media_cost"] == 'Sense') | (app_sum_org_1["Media_cost"] == 'svgmedia_int') | (app_sum_org_1["Media_cost"] == 'themobilyarabia_int') | (app_sum_org_1["Media_cost"] == 'Twitter') | (app_sum_org_1["Media_cost"] == 'tyroo_int') | (app_sum_org_1["Media_cost"] == 'uchuichuan_int') | (app_sum_org_1["Media_cost"] == 'vcommission_int') | (app_sum_org_1["Media_cost"] == 'vertozaff_int') | (app_sum_org_1["Media_cost"] == 'xaprio_int') | (app_sum_org_1["Media_cost"] == 'xyads_int') | (app_sum_org_1["Media_cost"] == 'yahoogemini_int'),'disb_amount']
	m_Aff_disbursal_amount_df.reset_index(inplace=True)
	#print (m_Aff_disbursal_amount_df)

	m_Aff_disbursal_total_amount_df = 0
	
	i = 0

	while (i <= m_Aff_disbursal_amount_df.last_valid_index()):
		m_Aff_disbursal_total_amount_df = m_Aff_disbursal_total_amount_df + m_Aff_disbursal_amount_df.iloc[i,1]
		i = i + 1

	print (m_Aff_disbursal_total_amount_df)


	print("########## Appnext APP COUNT ######################")
	
	#### Using df_mod
	m_Appnext_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Appnext','count_dis_Allstage']

	if (m_Appnext_app_count_df.empty == False):

		m_Appnext_app_count_df.columns = ['Count']
		m_Appnext_app_count_df.reset_index(inplace=True)
		m_Appnext_app_count = m_Appnext_app_count_df['Count'][0]
	

	else:
		m_Appnext_app_count = 0

	print(m_Appnext_app_count)

	print("########## Appnext Disbursal COUNT ######################")


	m_Appnext_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Appnext','disb_cust']

	if (m_Appnext_disbursal_count_df.empty == False):

		m_Appnext_disbursal_count_df.columns = ['Count']
		m_Appnext_disbursal_count_df.reset_index(inplace=True)
		m_Appnext_disbursal_count = m_Appnext_disbursal_count_df['Count'][0]
	

	else:
		m_Appnext_disbursal_count = 0

	print(m_Appnext_disbursal_count)


	print("########## Appnext Disbursal AMOUNT ######################")

	m_Appnext_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Appnext','disb_amount']

	if (m_Appnext_disbursal_amt_count_df.empty == False):

		m_Appnext_disbursal_amt_count_df.columns = ['Count']
		m_Appnext_disbursal_amt_count_df.reset_index(inplace=True)
		m_Appnext_disbursal_amt_count = m_Appnext_disbursal_amt_count_df['Count'][0]
	

	else:
		m_Appnext_disbursal_amt_count = 0


	print(m_Appnext_disbursal_amt_count)


	print("########## Madison = DBM APP COUNT ######################")

	df1 = df_mod.loc[(df_mod['Partner'] == '') | (df_mod['Partner'] == 'adapptmobi') | (df_mod['Partner'] == 'inmobiagen') | (df_mod['Partner'] == 'madison') | (df_mod['Partner'] == 'vserv'),:]

	if (df1.empty == False):
		
		#### Using df_mod
		
		app_sum_org_1 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		m_DBM_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','count_dis_Allstage']

		if (m_DBM_app_count_df.empty == False):

			m_DBM_app_count_df.columns = ['Count']
			m_DBM_app_count_df.reset_index(inplace=True)
			m_DBM_app_count = m_DBM_app_count_df['Count'][0]
		

		else:
			m_DBM_app_count = 0

		m_DBM_retarg_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','count_dis_Allstage']

		if (m_DBM_retarg_app_count_df.empty == False):

			m_DBM_retarg_app_count_df.columns = ['Count']
			m_DBM_retarg_app_count_df.reset_index(inplace=True)
			m_DBM_retarg_app_count = m_DBM_retarg_app_count_df['Count'][0]
		

		else:
			m_DBM_retarg_app_count = 0

		m_DBM_tot_app_count = m_DBM_app_count + m_DBM_retarg_app_count
		print(m_DBM_app_count,m_DBM_retarg_app_count,m_DBM_tot_app_count)

		print("########## Madison = DBM Disbursal COUNT ######################")


		m_DBM_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','disb_cust']

		if (m_DBM_disbursal_count_df.empty == False):

			m_DBM_disbursal_count_df.columns = ['Count']
			m_DBM_disbursal_count_df.reset_index(inplace=True)
			m_DBM_disbursal_count = m_DBM_disbursal_count_df['Count'][0]
		

		else:
			m_DBM_disbursal_count = 0

		m_DBM_retarg_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','disb_cust']

		if (m_DBM_retarg_disbursal_count_df.empty == False):

			m_DBM_retarg_disbursal_count_df.columns = ['Count']
			m_DBM_retarg_disbursal_count_df.reset_index(inplace=True)
			m_DBM_retarg_disbursal_count = m_DBM_retarg_disbursal_count_df['Count'][0]
		

		else:
			m_DBM_retarg_disbursal_count = 0

		m_DBM_tot_disbursal_count = m_DBM_disbursal_count + m_DBM_retarg_disbursal_count
		print(m_DBM_disbursal_count,m_DBM_retarg_disbursal_count,m_DBM_tot_disbursal_count)


		print("########## Madison = DBM Disbursal AMOUNT ######################")

		m_DBM_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','disb_amount']

		if (m_DBM_disbursal_amt_count_df.empty == False):

			m_DBM_disbursal_amt_count_df.columns = ['Count']
			m_DBM_disbursal_amt_count_df.reset_index(inplace=True)
			m_DBM_disbursal_amt_count = m_DBM_disbursal_amt_count_df['Count'][0]
		

		else:
			m_DBM_disbursal_amt_count = 0

		m_DBM_retarg_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','disb_amount']

		if (m_DBM_retarg_disbursal_amt_count_df.empty == False):

			m_DBM_retarg_disbursal_amt_count_df.columns = ['Count']
			m_DBM_retarg_disbursal_amt_count_df.reset_index(inplace=True)
			m_DBM_retarg_disbursal_amt_count = m_DBM_retarg_disbursal_amt_count_df['Count'][0]
		

		else:
			m_DBM_retarg_disbursal_amt_count = 0

		m_DBM_tot_disbursal_amt_count = m_DBM_disbursal_amt_count + m_DBM_retarg_disbursal_amt_count
		print(m_DBM_disbursal_amt_count,m_DBM_retarg_disbursal_amt_count,m_DBM_tot_disbursal_amt_count)
	
	else:
		m_DBM_tot_app_count = 0
		m_DBM_tot_disbursal_count = 0
		m_DBM_tot_disbursal_amt_count = 0

	print("########## Sokrati = DBM APP COUNT ######################")

	df1 = df_mod.loc[(df_mod['Partner'] == 'sokrati'),:]

	if (df1.empty == False):
		
		app_sum_org_1 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
		
		sk_DBM_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','count_dis_Allstage']

		if (sk_DBM_app_count_df.empty == False):

			sk_DBM_app_count_df.columns = ['Count']
			sk_DBM_app_count_df.reset_index(inplace=True)
			sk_DBM_app_count = sk_DBM_app_count_df['Count'][0]
		

		else:
			sk_DBM_app_count = 0

		sk_DBM_retarg_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','count_dis_Allstage']

		if (sk_DBM_retarg_app_count_df.empty == False):

			sk_DBM_retarg_app_count_df.columns = ['Count']
			sk_DBM_retarg_app_count_df.reset_index(inplace=True)
			sk_DBM_retarg_app_count = sk_DBM_retarg_app_count_df['Count'][0]
		

		else:
			sk_DBM_retarg_app_count = 0

		print(sk_DBM_app_count,sk_DBM_retarg_app_count)

		print("########## Sokrati = DBM Disbursal COUNT ######################")


		sk_DBM_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','disb_cust']

		if (sk_DBM_disbursal_count_df.empty == False):

			sk_DBM_disbursal_count_df.columns = ['Count']
			sk_DBM_disbursal_count_df.reset_index(inplace=True)
			sk_DBM_disbursal_count = sk_DBM_disbursal_count_df['Count'][0]
		

		else:
			sk_DBM_disbursal_count = 0

		sk_DBM_retarg_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','disb_cust']

		if (sk_DBM_retarg_disbursal_count_df.empty == False):

			sk_DBM_retarg_disbursal_count_df.columns = ['Count']
			sk_DBM_retarg_disbursal_count_df.reset_index(inplace=True)
			sk_DBM_retarg_disbursal_count = sk_DBM_retarg_disbursal_count_df['Count'][0]
		

		else:
			sk_DBM_retarg_disbursal_count = 0

		print(sk_DBM_disbursal_count,sk_DBM_retarg_disbursal_count)


		print("########## Sokrati = DBM Disbursal AMOUNT ######################")

		sk_DBM_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int','disb_amount']

		if (sk_DBM_disbursal_amt_count_df.empty == False):

			sk_DBM_disbursal_amt_count_df.columns = ['Count']
			sk_DBM_disbursal_amt_count_df.reset_index(inplace=True)
			sk_DBM_disbursal_amt_count = sk_DBM_disbursal_amt_count_df['Count'][0]
		

		else:
			sk_DBM_disbursal_amt_count = 0

		sk_DBM_retarg_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'doubleclick_int_Retargett','disb_amount']

		if (sk_DBM_retarg_disbursal_amt_count_df.empty == False):

			sk_DBM_retarg_disbursal_amt_count_df.columns = ['Count']
			sk_DBM_retarg_disbursal_amt_count_df.reset_index(inplace=True)
			sk_DBM_retarg_disbursal_amt_count = sk_DBM_retarg_disbursal_amt_count_df['Count'][0]
		

		else:
			sk_DBM_retarg_disbursal_amt_count = 0

		print(sk_DBM_disbursal_amt_count,sk_DBM_retarg_disbursal_amt_count)
	
	else:
		sk_DBM_app_count = 0
		sk_DBM_retarg_app_count = 0
		sk_DBM_disbursal_count = 0
		sk_DBM_retarg_disbursal_count = 0
		sk_DBM_disbursal_amt_count = 0
		sk_DBM_retarg_disbursal_amt_count = 0


	print("########## Inmobiagen-AppleSearch APP COUNT ######################")
	
	df1 = df_mod.loc[df_mod['Partner'] == 'inmobiagen',:]

	if (df1.empty == False):

		app_sum_org_1 = pd.pivot_table(df1, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
		app_sum_org_1.reset_index(inplace=True)
	

		m_AppleSearch_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Organic','count_dis_Allstage']

		if (m_AppleSearch_app_count_df.empty == False):

			m_AppleSearch_app_count_df.columns = ['Count']
			m_AppleSearch_app_count_df.reset_index(inplace=True)
			m_AppleSearch_app_count = m_AppleSearch_app_count_df['Count'][0]
		

		else:
			m_AppleSearch_app_count = 0

		print(m_AppleSearch_app_count)

		print("########## Inmobiagen-AppleSearch Disbursal COUNT ######################")


		m_AppleSearch_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Organic','disb_cust']

		if (m_AppleSearch_disbursal_count_df.empty == False):

			m_AppleSearch_disbursal_count_df.columns = ['Count']
			m_AppleSearch_disbursal_count_df.reset_index(inplace=True)
			m_AppleSearch_disbursal_count = m_AppleSearch_disbursal_count_df['Count'][0]
		

		else:
			m_AppleSearch_disbursal_count = 0

		print(m_AppleSearch_disbursal_count)


		print("########## Inmobiagen-AppleSearch Disbursal AMOUNT ######################")

		m_AppleSearch_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Organic','disb_amount']

		if (m_AppleSearch_disbursal_amt_count_df.empty == False):

			m_AppleSearch_disbursal_amt_count_df.columns = ['Count']
			m_AppleSearch_disbursal_amt_count_df.reset_index(inplace=True)
			m_AppleSearch_disbursal_amt_count = m_AppleSearch_disbursal_amt_count_df['Count'][0]
		

		else:
			m_AppleSearch_disbursal_amt_count = 0


		print(m_AppleSearch_disbursal_amt_count)
	
	else:
		m_AppleSearch_app_count = 0
		m_AppleSearch_disbursal_count = 0
		m_AppleSearch_disbursal_amt_count = 0


	print("########## MADISON TOTAL ######################")

	madi_total_app = m_google_app_count + m_fb_app_count + m_inmobi_tot_app_count + m_pocket_app_count + m_LeadBolt_app_count + m_DBM_tot_app_count + m_Appnext_app_count + m_fbads_app_count + m_AppleSearch_app_count

	madi_total_disbursal = m_google_disbursal_count + m_fb_disbursal_count + m_inmobi_tot_disbursal_count + m_pocket_disbursal_count + m_DBM_tot_disbursal_count + m_Appnext_disbursal_count + m_fbads_disbursal_count + m_AppleSearch_disbursal_count

	madi_total_disbursal_amt = m_google_disbursal_amount + m_fb_disbursal_amount + m_inmobi_tot_disbursal_amt_count + m_pocket_disbursal_amt_count + m_LeadBolt_disbursal_amt_count + m_DBM_tot_disbursal_amt_count + m_Appnext_disbursal_amt_count + m_fbads_disbursal_amount + m_AppleSearch_disbursal_amt_count


	print("########## Mediamath APP COUNT ######################")
	
	app_sum_org_1 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
	app_sum_org_1.reset_index(inplace=True)

	#### Using df_mod
	sk_mediamath_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'mediamath_int','count_dis_Allstage']

	if (sk_mediamath_app_count_df.empty == False):

		sk_mediamath_app_count_df.columns = ['Count']
		sk_mediamath_app_count_df.reset_index(inplace=True)
		sk_mediamath_app_count = sk_mediamath_app_count_df['Count'][0]
	

	else:
		sk_mediamath_app_count = 0

	print(sk_mediamath_app_count)

	print("########## Mediamath Disbursal COUNT ######################")


	sk_mediamath_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'mediamath_int','disb_cust']

	if (sk_mediamath_disbursal_count_df.empty == False):

		sk_mediamath_disbursal_count_df.columns = ['Count']
		sk_mediamath_disbursal_count_df.reset_index(inplace=True)
		sk_mediamath_disbursal_count = sk_mediamath_disbursal_count_df['Count'][0]
	

	else:
		sk_mediamath_disbursal_count = 0

	print(sk_mediamath_disbursal_count)


	print("########## Mediamath Disbursal AMOUNT ######################")

	sk_mediamath_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'mediamath_int','disb_amount']

	if (sk_mediamath_disbursal_amt_count_df.empty == False):

		sk_mediamath_disbursal_amt_count_df.columns = ['Count']
		sk_mediamath_disbursal_amt_count_df.reset_index(inplace=True)
		sk_mediamath_disbursal_amt_count = sk_mediamath_disbursal_amt_count_df['Count'][0]
	

	else:
		sk_mediamath_disbursal_amt_count = 0


	print(sk_mediamath_disbursal_amt_count)

	print("########## TikTok APP COUNT ######################")
	
	#### Using df_mod
	sk_TikTok_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'bytedanceglobal_int','count_dis_Allstage']

	if (sk_TikTok_app_count_df.empty == False):

		sk_TikTok_app_count_df.columns = ['Count']
		sk_TikTok_app_count_df.reset_index(inplace=True)
		sk_TikTok_app_count = sk_TikTok_app_count_df['Count'][0]
	

	else:
		sk_TikTok_app_count = 0

	print(sk_TikTok_app_count)

	print("########## TikTok Disbursal COUNT ######################")


	sk_TikTok_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'bytedanceglobal_int','disb_cust']

	if (sk_TikTok_disbursal_count_df.empty == False):

		sk_TikTok_disbursal_count_df.columns = ['Count']
		sk_TikTok_disbursal_count_df.reset_index(inplace=True)
		sk_TikTok_disbursal_count = sk_TikTok_disbursal_count_df['Count'][0]
	

	else:
		sk_TikTok_disbursal_count = 0

	print(sk_TikTok_disbursal_count)


	print("########## TikTok Disbursal AMOUNT ######################")

	sk_TikTok_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'bytedanceglobal_int','disb_amount']

	if (sk_TikTok_disbursal_amt_count_df.empty == False):

		sk_TikTok_disbursal_amt_count_df.columns = ['Count']
		sk_TikTok_disbursal_amt_count_df.reset_index(inplace=True)
		sk_TikTok_disbursal_amt_count = sk_TikTok_disbursal_amt_count_df['Count'][0]
	

	else:
		sk_TikTok_disbursal_amt_count = 0


	print(sk_TikTok_disbursal_amt_count)

	print("########## Sokrati Affilates (appsamurai_int) APP COUNT ######################")
	
	#### Using df_mod
	sk_Affilates_app_1_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'appsamurai_int','count_dis_Allstage']

	if (sk_Affilates_app_1_count_df.empty == False):

		sk_Affilates_app_1_count_df.columns = ['Count']
		sk_Affilates_app_1_count_df.reset_index(inplace=True)
		sk_Affilates_app_1_count = sk_Affilates_app_1_count_df['Count'][0]
	

	else:
		sk_Affilates_app_1_count = 0

	print(sk_Affilates_app_1_count)

	print("########## Sokrati Affilates (appsamurai_int) Disbursal COUNT ######################")


	sk_Affilates_disbursal_1_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'appsamurai_int','disb_cust']

	if (sk_Affilates_disbursal_1_count_df.empty == False):

		sk_Affilates_disbursal_1_count_df.columns = ['Count']
		sk_Affilates_disbursal_1_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_1_count = sk_Affilates_disbursal_1_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_1_count = 0

	print(sk_Affilates_disbursal_1_count)


	print("########## Sokrati Affilates (appsamurai_int) Disbursal AMOUNT ######################")

	sk_Affilates_disbursal_amt_1_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'appsamurai_int','disb_amount']

	if (sk_Affilates_disbursal_amt_1_count_df.empty == False):

		sk_Affilates_disbursal_amt_1_count_df.columns = ['Count']
		sk_Affilates_disbursal_amt_1_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_amt_1_count = sk_Affilates_disbursal_amt_1_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_amt_1_count = 0


	print(sk_Affilates_disbursal_amt_1_count)


	print("########## Sokrati Affilates (silverpush_int) APP COUNT ######################")
	
	#### Using df_mod
	sk_Affilates_app_2_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'silverpush_int','count_dis_Allstage']

	if (sk_Affilates_app_2_count_df.empty == False):

		sk_Affilates_app_2_count_df.columns = ['Count']
		sk_Affilates_app_2_count_df.reset_index(inplace=True)
		sk_Affilates_app_2_count = sk_Affilates_app_2_count_df['Count'][0]
	

	else:
		sk_Affilates_app_2_count = 0

	print(sk_Affilates_app_2_count)

	print("########## Sokrati Affilates (silverpush_int) Disbursal COUNT ######################")


	sk_Affilates_disbursal_2_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'silverpush_int','disb_cust']

	if (sk_Affilates_disbursal_2_count_df.empty == False):

		sk_Affilates_disbursal_2_count_df.columns = ['Count']
		sk_Affilates_disbursal_2_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_2_count = sk_Affilates_disbursal_2_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_2_count = 0

	print(sk_Affilates_disbursal_2_count)


	print("########## Sokrati Affilates (silverpush_int) Disbursal AMOUNT ######################")

	sk_Affilates_disbursal_amt_2_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'silverpush_int','disb_amount']

	if (sk_Affilates_disbursal_amt_2_count_df.empty == False):

		sk_Affilates_disbursal_amt_2_count_df.columns = ['Count']
		sk_Affilates_disbursal_amt_2_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_amt_2_count = sk_Affilates_disbursal_amt_2_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_amt_2_count = 0


	print(sk_Affilates_disbursal_amt_2_count)

	print("########## Sokrati Affilates (snapchat_int) APP COUNT ######################")
	
	#### Using df_mod
	sk_Affilates_app_3_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'snapchat_int','count_dis_Allstage']

	if (sk_Affilates_app_3_count_df.empty == False):

		sk_Affilates_app_3_count_df.columns = ['Count']
		sk_Affilates_app_3_count_df.reset_index(inplace=True)
		sk_Affilates_app_3_count = sk_Affilates_app_3_count_df['Count'][0]
	

	else:
		sk_Affilates_app_3_count = 0

	print(sk_Affilates_app_3_count)

	print("########## Sokrati Affilates (snapchat_int) Disbursal COUNT ######################")


	sk_Affilates_disbursal_3_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'snapchat_int','disb_cust']

	if (sk_Affilates_disbursal_3_count_df.empty == False):

		sk_Affilates_disbursal_3_count_df.columns = ['Count']
		sk_Affilates_disbursal_3_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_3_count = sk_Affilates_disbursal_3_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_3_count = 0

	print(sk_Affilates_disbursal_3_count)


	print("########## Sokrati Affilates (snapchat_int) Disbursal AMOUNT ######################")

	sk_Affilates_disbursal_amt_3_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'snapchat_int','disb_amount']

	if (sk_Affilates_disbursal_amt_3_count_df.empty == False):

		sk_Affilates_disbursal_amt_3_count_df.columns = ['Count']
		sk_Affilates_disbursal_amt_3_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_amt_3_count = sk_Affilates_disbursal_amt_3_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_amt_3_count = 0


	print(sk_Affilates_disbursal_amt_3_count)

	print("########## Sokrati Affilates (Optimize) APP COUNT ######################")
	
	#### Using df_mod
	sk_Affilates_app_4_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Optimize','count_dis_Allstage']

	if (sk_Affilates_app_4_count_df.empty == False):

		sk_Affilates_app_4_count_df.columns = ['Count']
		sk_Affilates_app_4_count_df.reset_index(inplace=True)
		sk_Affilates_app_4_count = sk_Affilates_app_4_count_df['Count'][0]
	

	else:
		sk_Affilates_app_4_count = 0

	print(sk_Affilates_app_4_count)

	print("########## Sokrati Affilates (Optimize) Disbursal COUNT ######################")


	sk_Affilates_disbursal_4_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Optimize','disb_cust']

	if (sk_Affilates_disbursal_4_count_df.empty == False):

		sk_Affilates_disbursal_4_count_df.columns = ['Count']
		sk_Affilates_disbursal_4_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_4_count = sk_Affilates_disbursal_4_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_4_count = 0

	print(sk_Affilates_disbursal_4_count)


	print("########## Sokrati Affilates (Optimize) Disbursal AMOUNT ######################")

	sk_Affilates_disbursal_amt_4_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Optimize','disb_amount']

	if (sk_Affilates_disbursal_amt_4_count_df.empty == False):

		sk_Affilates_disbursal_amt_4_count_df.columns = ['Count']
		sk_Affilates_disbursal_amt_4_count_df.reset_index(inplace=True)
		sk_Affilates_disbursal_amt_4_count = sk_Affilates_disbursal_amt_4_count_df['Count'][0]
	

	else:
		sk_Affilates_disbursal_amt_4_count = 0


	print(sk_Affilates_disbursal_amt_4_count)

	sk_Affilates_app_total_count = sk_Affilates_app_1_count + sk_Affilates_app_2_count + sk_Affilates_app_3_count + sk_Affilates_app_4_count
	sk_Affilates_disbursal_total_count = sk_Affilates_disbursal_1_count + sk_Affilates_disbursal_2_count + sk_Affilates_disbursal_3_count + sk_Affilates_disbursal_4_count
	sk_Affilates_disbursal_amt_total_count = sk_Affilates_disbursal_amt_1_count + sk_Affilates_disbursal_amt_2_count + sk_Affilates_disbursal_amt_3_count + sk_Affilates_disbursal_amt_4_count
	
	print (sk_Affilates_app_total_count,sk_Affilates_disbursal_total_count,sk_Affilates_disbursal_amt_total_count)


	print("########## SOKRATI TOTAL ######################")

	sok_total_app = sk_fb_app_tot_count + sk_mediamath_app_count + sk_DBM_app_count + sk_google_app_count + sk_TikTok_app_count + sk_Affilates_app_total_count + sk_DBM_retarg_app_count

	sok_total_disbursal = sk_fb_disbursal_tot_count + sk_mediamath_disbursal_count + sk_DBM_disbursal_count + sk_google_disbursal_count + sk_TikTok_disbursal_count + sk_Affilates_disbursal_total_count +  sk_DBM_retarg_disbursal_count

	sok_total_disbursal_amt = sk_fb_disbursal_tot_amount + sk_mediamath_disbursal_amt_count + sk_DBM_disbursal_amt_count + sk_google_disbursal_amount + sk_TikTok_disbursal_amt_count + sk_Affilates_disbursal_amt_total_count + sk_DBM_retarg_disbursal_amt_count


	print("########## seventynine APP COUNT ######################")
	
	#### Using df_mod
	seventynine_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'seventyninemobi_int','count_dis_Allstage']

	if (seventynine_app_count_df.empty == False):

		seventynine_app_count_df.columns = ['Count']
		seventynine_app_count_df.reset_index(inplace=True)
		seventynine_app_count = seventynine_app_count_df['Count'][0]
	

	else:
		seventynine_app_count = 0

	print(seventynine_app_count)

	print("########## seventynine Disbursal COUNT ######################")


	seventynine_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'seventyninemobi_int','disb_cust']

	if (seventynine_disbursal_count_df.empty == False):

		seventynine_disbursal_count_df.columns = ['Count']
		seventynine_disbursal_count_df.reset_index(inplace=True)
		seventynine_disbursal_count = seventynine_disbursal_count_df['Count'][0]
	

	else:
		seventynine_disbursal_count = 0

	print(seventynine_disbursal_count)


	print("########## seventynine Disbursal AMOUNT ######################")

	seventynine_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'seventyninemobi_int','disb_amount']

	if (seventynine_disbursal_amt_count_df.empty == False):

		seventynine_disbursal_amt_count_df.columns = ['Count']
		seventynine_disbursal_amt_count_df.reset_index(inplace=True)
		seventynine_disbursal_amt_count = seventynine_disbursal_amt_count_df['Count'][0]
	

	else:
		seventynine_disbursal_amt_count = 0


	print(seventynine_disbursal_amt_count)


	print("########## BuddyLoan APP COUNT ######################")
	
	#### Using df_mod
	BuddyLoan_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Buddy Loan','count_dis_Allstage']

	if (BuddyLoan_app_count_df.empty == False):

		BuddyLoan_app_count_df.columns = ['Count']
		BuddyLoan_app_count_df.reset_index(inplace=True)
		BuddyLoan_app_count = BuddyLoan_app_count_df['Count'][0]
	

	else:
		BuddyLoan_app_count = 0

	print(BuddyLoan_app_count)

	print("########## BuddyLoan Disbursal COUNT ######################")


	BuddyLoan_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Buddy Loan','disb_cust']

	if (BuddyLoan_disbursal_count_df.empty == False):

		BuddyLoan_disbursal_count_df.columns = ['Count']
		BuddyLoan_disbursal_count_df.reset_index(inplace=True)
		BuddyLoan_disbursal_count = BuddyLoan_disbursal_count_df['Count'][0]
	

	else:
		BuddyLoan_disbursal_count = 0

	print(BuddyLoan_disbursal_count)


	print("########## BuddyLoan Disbursal AMOUNT ######################")

	BuddyLoan_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'Buddy Loan','disb_amount']

	if (BuddyLoan_disbursal_amt_count_df.empty == False):

		BuddyLoan_disbursal_amt_count_df.columns = ['Count']
		BuddyLoan_disbursal_amt_count_df.reset_index(inplace=True)
		BuddyLoan_disbursal_amt_count = BuddyLoan_disbursal_amt_count_df['Count'][0]
	

	else:
		BuddyLoan_disbursal_amt_count = 0


	print(BuddyLoan_disbursal_amt_count)

	print("########## Netcore APP COUNT ######################")
		
	#### Using df_mod
	Netcore_app_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'netcore_int','count_dis_Allstage']

	if (Netcore_app_count_df.empty == False):

		Netcore_app_count_df.columns = ['Count']
		Netcore_app_count_df.reset_index(inplace=True)
		Netcore_app_count = Netcore_app_count_df['Count'][0]
	

	else:
		Netcore_app_count = 0

	print(Netcore_app_count)

	print("########## Netcore Disbursal COUNT ######################")


	Netcore_disbursal_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'netcore_int','disb_cust']

	if (Netcore_disbursal_count_df.empty == False):

		Netcore_disbursal_count_df.columns = ['Count']
		Netcore_disbursal_count_df.reset_index(inplace=True)
		Netcore_disbursal_count = Netcore_disbursal_count_df['Count'][0]
	

	else:
		Netcore_disbursal_count = 0

	print(Netcore_disbursal_count)


	print("########## Netcore Disbursal AMOUNT ######################")

	Netcore_disbursal_amt_count_df = app_sum_org_1.loc[app_sum_org_1["Media_cost"] == 'netcore_int','disb_amount']

	if (Netcore_disbursal_amt_count_df.empty == False):

		Netcore_disbursal_amt_count_df.columns = ['Count']
		Netcore_disbursal_amt_count_df.reset_index(inplace=True)
		Netcore_disbursal_amt_count = Netcore_disbursal_amt_count_df['Count'][0]
	

	else:
		Netcore_disbursal_amt_count = 0


	print(Netcore_disbursal_amt_count)


	print("########## Others Inorganic = APP COUNT ######################")
	#### Using df_mod
	
	other_inorg_app_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'CD-worldcup-1106') | (app_sum_org_1["Media_cost"] == 'FB_Group_S') | (app_sum_org_1["Media_cost"] == 'HA-Aug-SMS') | (app_sum_org_1["Media_cost"] == 'HA-LandingPage') | (app_sum_org_1["Media_cost"] == 'HA-SegD2,D3-0907') | (app_sum_org_1["Media_cost"] == 'P.1-SMS-0709') | (app_sum_org_1["Media_cost"] == 'PA-0108') | (app_sum_org_1["Media_cost"] == 'PA-NDND-0708') | (app_sum_org_1["Media_cost"] == 'PA-P3-0708') | (app_sum_org_1["Media_cost"] == 'PA-P4-09-08-19') | (app_sum_org_1["Media_cost"] == 'PA-Responders-SMS2') | (app_sum_org_1["Media_cost"] == 'Pre-Approved - LP') | (app_sum_org_1["Media_cost"] == 'Pre-Approved-0307') | (app_sum_org_1["Media_cost"] == 'Seg F-1608') | (app_sum_org_1["Media_cost"] == 'SegJ-1308') | (app_sum_org_1["Media_cost"] == 'SEGJ1-SMS-2706') | (app_sum_org_1["Media_cost"] == 'SegK-Mobileregistered') | (app_sum_org_1["Media_cost"] == 'SMS-HA-0309') | (app_sum_org_1["Media_cost"] == 'SMS-Topup-0309') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051,') | (app_sum_org_1["Media_cost"] == 'TopUp-IntroDP-1809') | (app_sum_org_1["Media_cost"] == 'TopUp-LandingPage-2506') | (app_sum_org_1["Media_cost"] == 'TopUpLoan_NewDatabase250') | (app_sum_org_1["Media_cost"] == 'Top-Up-SMS-090819') | (app_sum_org_1["Media_cost"] == 'TP-SegD-2106') | (app_sum_org_1["Media_cost"] == 'TP-SegDregional-2506') | (app_sum_org_1["Media_cost"] == 'TP-SMS-1207') | (app_sum_org_1["Media_cost"] == 'TP-worldcup-1006') | (app_sum_org_1["Media_cost"] == 'Yaarii2Dhani') | (app_sum_org_1["Media_cost"] == 'Youtube') | (app_sum_org_1["Media_cost"] == 'YT-Pre-Approved'),'count_dis_Allstage']
	other_inorg_app_count_df.reset_index(inplace=True)
	print (other_inorg_app_count_df)

	other_inorg_app_total_count_df = 0
	
	i = 0

	while (i <= other_inorg_app_count_df.last_valid_index()):
		other_inorg_app_total_count_df = other_inorg_app_total_count_df + other_inorg_app_count_df.iloc[i,1]
		i = i + 1

	print (other_inorg_app_total_count_df)



	print("########## Others Inorganic =  Disbursal COUNT ######################")

	other_inorg_disbursal_count_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'CD-worldcup-1106') | (app_sum_org_1["Media_cost"] == 'FB_Group_S') | (app_sum_org_1["Media_cost"] == 'HA-Aug-SMS') | (app_sum_org_1["Media_cost"] == 'HA-LandingPage') | (app_sum_org_1["Media_cost"] == 'HA-SegD2,D3-0907') | (app_sum_org_1["Media_cost"] == 'P.1-SMS-0709') | (app_sum_org_1["Media_cost"] == 'PA-0108') | (app_sum_org_1["Media_cost"] == 'PA-NDND-0708') | (app_sum_org_1["Media_cost"] == 'PA-P3-0708') | (app_sum_org_1["Media_cost"] == 'PA-P4-09-08-19') | (app_sum_org_1["Media_cost"] == 'PA-Responders-SMS2') | (app_sum_org_1["Media_cost"] == 'Pre-Approved - LP') | (app_sum_org_1["Media_cost"] == 'Pre-Approved-0307') | (app_sum_org_1["Media_cost"] == 'Seg F-1608') | (app_sum_org_1["Media_cost"] == 'SegJ-1308') | (app_sum_org_1["Media_cost"] == 'SEGJ1-SMS-2706') | (app_sum_org_1["Media_cost"] == 'SegK-Mobileregistered') | (app_sum_org_1["Media_cost"] == 'SMS-HA-0309') | (app_sum_org_1["Media_cost"] == 'SMS-Topup-0309') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051,') | (app_sum_org_1["Media_cost"] == 'TopUp-IntroDP-1809') | (app_sum_org_1["Media_cost"] == 'TopUp-LandingPage-2506') | (app_sum_org_1["Media_cost"] == 'TopUpLoan_NewDatabase250') | (app_sum_org_1["Media_cost"] == 'Top-Up-SMS-090819') | (app_sum_org_1["Media_cost"] == 'TP-SegD-2106') | (app_sum_org_1["Media_cost"] == 'TP-SegDregional-2506') | (app_sum_org_1["Media_cost"] == 'TP-SMS-1207') | (app_sum_org_1["Media_cost"] == 'TP-worldcup-1006') | (app_sum_org_1["Media_cost"] == 'Yaarii2Dhani') | (app_sum_org_1["Media_cost"] == 'Youtube') | (app_sum_org_1["Media_cost"] == 'YT-Pre-Approved'),'disb_cust']
	other_inorg_disbursal_count_df.reset_index(inplace=True)
	#print (other_inorg_disbursal_count_df)

	other_inorg_disbursal_total_count_df = 0
	
	i = 0

	while (i <= other_inorg_disbursal_count_df.last_valid_index()):
		other_inorg_disbursal_total_count_df = other_inorg_disbursal_total_count_df + other_inorg_disbursal_count_df.iloc[i,1]
		i = i + 1

	print (other_inorg_disbursal_total_count_df)

	print("########## Others Inorganic =  Disbursal AMOUNT ######################")

	other_inorg_disbursal_amount_df = app_sum_org_1.loc[(app_sum_org_1["Media_cost"] == 'CD-worldcup-1106') | (app_sum_org_1["Media_cost"] == 'FB_Group_S') | (app_sum_org_1["Media_cost"] == 'HA-Aug-SMS') | (app_sum_org_1["Media_cost"] == 'HA-LandingPage') | (app_sum_org_1["Media_cost"] == 'HA-SegD2,D3-0907') | (app_sum_org_1["Media_cost"] == 'P.1-SMS-0709') | (app_sum_org_1["Media_cost"] == 'PA-0108') | (app_sum_org_1["Media_cost"] == 'PA-NDND-0708') | (app_sum_org_1["Media_cost"] == 'PA-P3-0708') | (app_sum_org_1["Media_cost"] == 'PA-P4-09-08-19') | (app_sum_org_1["Media_cost"] == 'PA-Responders-SMS2') | (app_sum_org_1["Media_cost"] == 'Pre-Approved - LP') | (app_sum_org_1["Media_cost"] == 'Pre-Approved-0307') | (app_sum_org_1["Media_cost"] == 'Seg F-1608') | (app_sum_org_1["Media_cost"] == 'SegJ-1308') | (app_sum_org_1["Media_cost"] == 'SEGJ1-SMS-2706') | (app_sum_org_1["Media_cost"] == 'SegK-Mobileregistered') | (app_sum_org_1["Media_cost"] == 'SMS-HA-0309') | (app_sum_org_1["Media_cost"] == 'SMS-Topup-0309') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051') | (app_sum_org_1["Media_cost"] == 'Top Up Campaign SG21051,') | (app_sum_org_1["Media_cost"] == 'TopUp-IntroDP-1809') | (app_sum_org_1["Media_cost"] == 'TopUp-LandingPage-2506') | (app_sum_org_1["Media_cost"] == 'TopUpLoan_NewDatabase250') | (app_sum_org_1["Media_cost"] == 'Top-Up-SMS-090819') | (app_sum_org_1["Media_cost"] == 'TP-SegD-2106') | (app_sum_org_1["Media_cost"] == 'TP-SegDregional-2506') | (app_sum_org_1["Media_cost"] == 'TP-SMS-1207') | (app_sum_org_1["Media_cost"] == 'TP-worldcup-1006') | (app_sum_org_1["Media_cost"] == 'Yaarii2Dhani') | (app_sum_org_1["Media_cost"] == 'Youtube') | (app_sum_org_1["Media_cost"] == 'YT-Pre-Approved'),'disb_amount']
	other_inorg_disbursal_amount_df.reset_index(inplace=True)
	#print (other_inorg_disbursal_amount_df)

	other_inorg_disbursal_total_amount_df = 0
	
	i = 0

	while (i <= other_inorg_disbursal_amount_df.last_valid_index()):
		other_inorg_disbursal_total_amount_df = other_inorg_disbursal_total_amount_df + other_inorg_disbursal_amount_df.iloc[i,1]
		i = i + 1

	print (other_inorg_disbursal_total_amount_df)


	others_app_count = seventynine_app_count + BuddyLoan_app_count + Netcore_app_count + other_inorg_app_total_count_df
	others_disbursal_count = seventynine_disbursal_count + BuddyLoan_disbursal_count + Netcore_disbursal_count + other_inorg_disbursal_total_count_df
	others_disbursal_amt_count = seventynine_disbursal_amt_count + BuddyLoan_disbursal_amt_count + Netcore_disbursal_amt_count + other_inorg_disbursal_total_amount_df

	app_count_total = app_blanks[0] + madi_total_app + sok_total_app + others_app_count
	disbursal_count_total = disbursal_blanks[0] + madi_total_disbursal +  sok_total_disbursal + others_disbursal_count
	disbursal_amount_total = disbursal_amt_blanks[0] + madi_total_disbursal_amt + sok_total_disbursal_amt + others_disbursal_amt_count


	
	print("########## WEB Application sum ORGANIC ######################")


	web_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_app_sum_org_2.reset_index(inplace=True)

	web_app_blanks = list(web_app_sum_org_2.loc[web_app_sum_org_2["Media_cost"] == 'WEB','Total'])
	
	if (web_app_blanks !=[]):
		web_app_blanks = [x for x in web_app_blanks if ~np.isnan(x)]
		web_app_blanks.append(0)
	else:
		web_app_blanks.append(0)

	print (web_app_blanks[0])


	print("########## WEB Disbursal Count ORGANIC ######################")

	web_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_disbursal_sum_org_2.reset_index(inplace=True)

	web_disbursal_blanks = list(web_disbursal_sum_org_2.loc[web_disbursal_sum_org_2["Media_cost"] == 'WEB','Total'])

	if (web_disbursal_blanks !=[]):
		web_disbursal_blanks = [x for x in web_disbursal_blanks if ~np.isnan(x)]
		web_disbursal_blanks.append(0)
	else:
		web_disbursal_blanks.append(0)

	print (web_disbursal_blanks[0])

	print("########## WEB Disbursal AMOUNT ORGANIC ######################")

	web_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_disbursal_amount_org_2.reset_index(inplace=True)

	web_disbursal_amt_blanks = list(web_disbursal_amount_org_2.loc[web_disbursal_amount_org_2["Media_cost"] == 'WEB','Total'])
	
	if (web_disbursal_amt_blanks !=[]):
		web_disbursal_amt_blanks = [x for x in web_disbursal_amt_blanks if ~np.isnan(x)]
		web_disbursal_amt_blanks.append(0)
	else:
		web_disbursal_amt_blanks.append(0)

	print (web_disbursal_amt_blanks[0])


	print("########## WEB FACEBOOK Application sum ORGANIC ######################")


	web_fb_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_fb_app_sum_org_2.reset_index(inplace=True)

	web_fb_app_blanks = list(web_fb_app_sum_org_2.loc[web_fb_app_sum_org_2["Media_cost"] == 'Web Facebook','Total'])
	
	if (web_fb_app_blanks !=[]):
		web_fb_app_blanks = [x for x in web_fb_app_blanks if ~np.isnan(x)]
		web_fb_app_blanks.append(0)
	else:
		web_fb_app_blanks.append(0)

	print (web_fb_app_blanks[0])


	print("########## WEB FACEBOOK Disbursal Count ORGANIC ######################")

	web_fb_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_fb_disbursal_sum_org_2.reset_index(inplace=True)

	web_fb_disbursal_blanks = list(web_fb_disbursal_sum_org_2.loc[web_fb_disbursal_sum_org_2["Media_cost"] == 'Web Facebook','Total'])

	if (web_fb_disbursal_blanks !=[]):
		web_fb_disbursal_blanks = [x for x in web_fb_disbursal_blanks if ~np.isnan(x)]
		web_fb_disbursal_blanks.append(0)
	else:
		web_fb_disbursal_blanks.append(0)

	print (web_fb_disbursal_blanks[0])

	print("########## WEB FACEBOOK Disbursal AMOUNT ORGANIC ######################")

	web_fb_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_fb_disbursal_amount_org_2.reset_index(inplace=True)

	web_fb_disbursal_amt_blanks = list(web_fb_disbursal_amount_org_2.loc[web_fb_disbursal_amount_org_2["Media_cost"] == 'Web Facebook','Total'])
	
	if (web_fb_disbursal_amt_blanks !=[]):
		web_fb_disbursal_amt_blanks = [x for x in web_fb_disbursal_amt_blanks if ~np.isnan(x)]
		web_fb_disbursal_amt_blanks.append(0)
	else:
		web_fb_disbursal_amt_blanks.append(0)

	print (web_fb_disbursal_amt_blanks[0])


	print("########## Web Google Application sum ORGANIC ######################")


	web_google_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_google_app_sum_org_2.reset_index(inplace=True)

	web_google_app_blanks = list(web_google_app_sum_org_2.loc[web_google_app_sum_org_2["Media_cost"] == 'Web Google','Total'])
	
	if (web_google_app_blanks !=[]):
		web_google_app_blanks = [x for x in web_google_app_blanks if ~np.isnan(x)]
		web_google_app_blanks.append(0)
	else:
		web_google_app_blanks.append(0)

	print (web_google_app_blanks[0])


	print("########## Web Google Disbursal Count ORGANIC ######################")

	web_google_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_google_disbursal_sum_org_2.reset_index(inplace=True)

	web_google_disbursal_blanks = list(web_google_disbursal_sum_org_2.loc[web_google_disbursal_sum_org_2["Media_cost"] == 'Web Google','Total'])

	if (web_google_disbursal_blanks !=[]):
		web_google_disbursal_blanks = [x for x in web_google_disbursal_blanks if ~np.isnan(x)]
		web_google_disbursal_blanks.append(0)
	else:
		web_google_disbursal_blanks.append(0)

	print (web_google_disbursal_blanks[0])

	print("########## Web Google Disbursal AMOUNT ORGANIC ######################")

	web_google_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_google_disbursal_amount_org_2.reset_index(inplace=True)

	web_google_disbursal_amt_blanks = list(web_google_disbursal_amount_org_2.loc[web_google_disbursal_amount_org_2["Media_cost"] == 'Web Google','Total'])
	
	if (web_google_disbursal_amt_blanks !=[]):
		web_google_disbursal_amt_blanks = [x for x in web_google_disbursal_amt_blanks if ~np.isnan(x)]
		web_google_disbursal_amt_blanks.append(0)
	else:
		web_google_disbursal_amt_blanks.append(0)

	print (web_google_disbursal_amt_blanks[0])

	print("########## Web ValueLeaf Application sum ORGANIC ######################")


	web_vlf_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_vlf_app_sum_org_2.reset_index(inplace=True)

	web_vlf_app_blanks = list(web_vlf_app_sum_org_2.loc[web_vlf_app_sum_org_2["Media_cost"] == 'Web ValueLeaf','Total'])
	
	if (web_vlf_app_blanks !=[]):
		web_vlf_app_blanks = [x for x in web_vlf_app_blanks if ~np.isnan(x)]
		web_vlf_app_blanks.append(0)
	else:
		web_vlf_app_blanks.append(0)

	print (web_vlf_app_blanks[0])


	print("########## Web ValueLeaf Disbursal Count ORGANIC ######################")

	web_vlf_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_vlf_disbursal_sum_org_2.reset_index(inplace=True)

	web_vlf_disbursal_blanks = list(web_vlf_disbursal_sum_org_2.loc[web_vlf_disbursal_sum_org_2["Media_cost"] == 'Web ValueLeaf','Total'])

	if (web_vlf_disbursal_blanks !=[]):
		web_vlf_disbursal_blanks = [x for x in web_vlf_disbursal_blanks if ~np.isnan(x)]
		web_vlf_disbursal_blanks.append(0)
	else:
		web_vlf_disbursal_blanks.append(0)

	print (web_vlf_disbursal_blanks[0])

	print("########## Web ValueLeaf Disbursal AMOUNT ORGANIC ######################")

	web_vlf_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_vlf_disbursal_amount_org_2.reset_index(inplace=True)

	web_vlf_disbursal_amt_blanks = list(web_vlf_disbursal_amount_org_2.loc[web_vlf_disbursal_amount_org_2["Media_cost"] == 'Web ValueLeaf','Total'])
	
	if (web_vlf_disbursal_amt_blanks !=[]):
		web_vlf_disbursal_amt_blanks = [x for x in web_vlf_disbursal_amt_blanks if ~np.isnan(x)]
		web_vlf_disbursal_amt_blanks.append(0)
	else:
		web_vlf_disbursal_amt_blanks.append(0)

	print (web_vlf_disbursal_amt_blanks[0])


	print("########## WEB2ONLINE Application sum ORGANIC ######################")


	web_2_onl_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_2_onl_app_sum_org_2.reset_index(inplace=True)

	web_2_onl_app_blanks = list(web_2_onl_app_sum_org_2.loc[web_2_onl_app_sum_org_2["Media_cost"] == 'WEB2ONLINE','Total'])
	
	if (web_2_onl_app_blanks !=[]):
		web_2_onl_app_blanks = [x for x in web_2_onl_app_blanks if ~np.isnan(x)]
		web_2_onl_app_blanks.append(0)
	else:
		web_2_onl_app_blanks.append(0)

	print (web_2_onl_app_blanks[0])


	print("########## WEB2ONLINE Disbursal Count ORGANIC ######################")

	web_2_onl_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_2_onl_disbursal_sum_org_2.reset_index(inplace=True)

	web_2_onl_disbursal_blanks = list(web_2_onl_disbursal_sum_org_2.loc[web_2_onl_disbursal_sum_org_2["Media_cost"] == 'WEB2ONLINE','Total'])

	if (web_2_onl_disbursal_blanks !=[]):
		web_2_onl_disbursal_blanks = [x for x in web_2_onl_disbursal_blanks if ~np.isnan(x)]
		web_2_onl_disbursal_blanks.append(0)
	else:
		web_2_onl_disbursal_blanks.append(0)

	print (web_2_onl_disbursal_blanks[0])

	print("########## WEB2ONLINE Disbursal AMOUNT ORGANIC ######################")

	web_2_onl_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_2_onl_disbursal_amount_org_2.reset_index(inplace=True)

	web_2_onl_disbursal_amt_blanks = list(web_2_onl_disbursal_amount_org_2.loc[web_2_onl_disbursal_amount_org_2["Media_cost"] == 'WEB2ONLINE','Total'])
	
	if (web_2_onl_disbursal_amt_blanks !=[]):
		web_2_onl_disbursal_amt_blanks = [x for x in web_2_onl_disbursal_amt_blanks if ~np.isnan(x)]
		web_2_onl_disbursal_amt_blanks.append(0)
	else:
		web_2_onl_disbursal_amt_blanks.append(0)

	print (web_2_onl_disbursal_amt_blanks[0])


	print("########## WEB Netcore Application sum ORGANIC ######################")


	web_netcore_app_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='count_dis_Allstage', aggfunc='sum')
	web_netcore_app_sum_org_2.reset_index(inplace=True)

	web_netcore_app_blanks = list(web_netcore_app_sum_org_2.loc[web_netcore_app_sum_org_2["Media_cost"] == 'WEB Netcore','Total'])
	
	if (web_netcore_app_blanks !=[]):
		web_netcore_app_blanks = [x for x in web_netcore_app_blanks if ~np.isnan(x)]
		web_netcore_app_blanks.append(0)
	else:
		web_netcore_app_blanks.append(0)

	print (web_netcore_app_blanks[0])


	print("########## WEB Netcore Disbursal Count ORGANIC ######################")

	web_netcore_disbursal_sum_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_cust', aggfunc='sum')
	web_netcore_disbursal_sum_org_2.reset_index(inplace=True)

	web_netcore_disbursal_blanks = list(web_netcore_disbursal_sum_org_2.loc[web_netcore_disbursal_sum_org_2["Media_cost"] == 'WEB Netcore','Total'])

	if (web_netcore_disbursal_blanks !=[]):
		web_netcore_disbursal_blanks = [x for x in web_netcore_disbursal_blanks if ~np.isnan(x)]
		web_netcore_disbursal_blanks.append(0)
	else:
		web_netcore_disbursal_blanks.append(0)

	print (web_netcore_disbursal_blanks[0])

	print("########## WEB Netcore Disbursal AMOUNT ORGANIC ######################")

	web_netcore_disbursal_amount_org_2 = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date',  margins=True, margins_name='Total', values='disb_amount', aggfunc='sum')
	web_netcore_disbursal_amount_org_2.reset_index(inplace=True)

	web_netcore_disbursal_amt_blanks = list(web_netcore_disbursal_amount_org_2.loc[web_netcore_disbursal_amount_org_2["Media_cost"] == 'WEB Netcore','Total'])


	if(web_netcore_disbursal_amt_blanks !=[]):
		web_netcore_disbursal_amt_blanks = [x for x in web_netcore_disbursal_amt_blanks if ~np.isnan(x)]
		web_netcore_disbursal_amt_blanks.append(0)
	else:
		web_netcore_disbursal_amt_blanks.append(0)

	print (web_netcore_disbursal_amt_blanks[0])
	
	print("########## WEB Affiliates ######################")

	df = pd.pivot_table(df_mod, index = 'Media_cost', columns='Created date', values=['count_dis_Allstage','disb_cust','disb_amount'], aggfunc='sum')
	df.reset_index(inplace=True)
	#df.to_csv("df.csv")
	exclusion = ['WEB','Web Facebook','Web Google','WEB Netcore','Web ValueLeaf','WEB2ONLINE']


	web_aff_app_count_sum = 0
	web_aff_disbursal_count_sum = 0
	web_aff_disbursal_amount_sum = 0

	
	for x in df["Media_cost"]:
	
		if ((x.startswith('WEB') | x.startswith('Web')) and (x not in exclusion)):
			print (x)
			app_count_df = df.loc[df["Media_cost"] == x,'count_dis_Allstage']
			app_count_df.columns = ['Count']
			app_count_df.reset_index(inplace=True)
			app_count = app_count_df['Count'][0]
			
			disbursal_count_df = df.loc[df["Media_cost"] == x,'disb_cust']
			disbursal_count_df.columns = ['Count']
			disbursal_count_df.reset_index(inplace=True)
			disbursal_count = disbursal_count_df['Count'][0]
		
			disbursal_amount_df = df.loc[df["Media_cost"] == x,'disb_amount']
			disbursal_amount_df.columns = ['Count']
			disbursal_amount_df.reset_index(inplace=True)
			disbursal_amount = disbursal_amount_df['Count'][0]

			web_aff_app_count_sum = web_aff_app_count_sum + app_count
			web_aff_disbursal_count_sum = web_aff_disbursal_count_sum + disbursal_count
			web_aff_disbursal_amount_sum = web_aff_disbursal_amount_sum + disbursal_amount


	print (web_aff_app_count_sum,web_aff_disbursal_count_sum,web_aff_disbursal_amount_sum)

	print("########## WEB Affiliates ENDS ######################")
	
	web_app_tot = web_app_blanks[0] + web_fb_app_blanks[0] + web_google_app_blanks[0] + web_aff_app_count_sum + web_vlf_app_blanks[0] + web_2_onl_app_blanks[0] + web_netcore_app_blanks[0]

	web_disbursal_tot = web_disbursal_blanks[0] + web_fb_disbursal_blanks[0] + web_google_disbursal_blanks[0] + web_aff_disbursal_count_sum + web_vlf_disbursal_blanks[0] + web_2_onl_disbursal_blanks[0] + web_netcore_disbursal_blanks[0]

	web_disbursal_amt_tot = web_disbursal_amt_blanks[0] + web_fb_disbursal_amt_blanks[0] + web_google_disbursal_amt_blanks[0] + web_aff_disbursal_amount_sum + web_vlf_disbursal_amt_blanks[0] + web_2_onl_disbursal_amt_blanks[0] + web_netcore_disbursal_amt_blanks[0]

	print (web_app_tot,web_disbursal_tot,web_disbursal_amt_tot)


	market_tot_app = app_count_total + web_app_tot
	market_tot_disbursal = disbursal_count_total + web_disbursal_tot
	market_tot_disbursal_amt = disbursal_amount_total + web_disbursal_amt_tot

	print (market_tot_app,market_tot_disbursal,market_tot_disbursal_amt)

	data_app = []
	data_disbursal = []
	data_disbursal_amt =[]

	data_app.append(app_blanks[0])
	data_app.append(m_google_app_count)
	data_app.append(m_fb_app_count)
	data_app.append(m_inmobi_tot_app_count)
	data_app.append(m_pocket_app_count)
	data_app.append(m_LeadBolt_app_count)
	data_app.append(m_Aff_app_total_count_df)
	data_app.append(m_DBM_tot_app_count)
	data_app.append(m_Appnext_app_count)
	data_app.append(m_fbads_app_count)
	data_app.append(m_AppleSearch_app_count)
	data_app.append(madi_total_app)
	data_app.append(sk_fb_app_tot_count)
	data_app.append(sk_mediamath_app_count)
	data_app.append(sk_DBM_app_count)
	data_app.append(sk_google_app_count)
	data_app.append(sk_TikTok_app_count)
	data_app.append(sk_Affilates_app_total_count)
	data_app.append(sk_DBM_retarg_app_count)
	data_app.append(sok_total_app)
	data_app.append(seventynine_app_count)
	data_app.append(BuddyLoan_app_count)
	data_app.append(Netcore_app_count)
	data_app.append(other_inorg_app_total_count_df)
	data_app.append(app_count_total)
	data_app.append("")
	data_app.append(web_app_blanks[0])
	data_app.append(web_fb_app_blanks[0])
	data_app.append(web_google_app_blanks[0])
	data_app.append(web_aff_app_count_sum)
	data_app.append(web_vlf_app_blanks[0])
	data_app.append(web_2_onl_app_blanks[0])
	data_app.append(web_netcore_app_blanks[0])
	data_app.append(web_app_tot)
	data_app.append("")
	data_app.append(market_tot_app)
	print data_app

	
	data_disbursal.append(disbursal_blanks[0])
	data_disbursal.append(m_google_disbursal_count)
	data_disbursal.append(m_fb_disbursal_count)
	data_disbursal.append(m_inmobi_tot_disbursal_count)
	data_disbursal.append(m_pocket_disbursal_count)
	data_disbursal.append(m_LeadBolt_disbursal_count)
	data_disbursal.append(m_Aff_disbursal_total_count_df)
	data_disbursal.append(m_DBM_tot_disbursal_count)
	data_disbursal.append(m_Appnext_disbursal_count)
	data_disbursal.append(m_fbads_disbursal_count)
	data_disbursal.append(m_AppleSearch_disbursal_count)
	data_disbursal.append(madi_total_disbursal)
	data_disbursal.append(sk_fb_disbursal_tot_count)
	data_disbursal.append(sk_mediamath_disbursal_count)
	data_disbursal.append(sk_DBM_disbursal_count)
	data_disbursal.append(sk_google_disbursal_count)
	data_disbursal.append(sk_TikTok_disbursal_count)
	data_disbursal.append(sk_Affilates_disbursal_total_count)
	data_disbursal.append(sk_DBM_retarg_disbursal_count)
	data_disbursal.append(sok_total_disbursal)
	data_disbursal.append(seventynine_disbursal_count)
	data_disbursal.append(BuddyLoan_disbursal_count)
	data_disbursal.append(Netcore_disbursal_count)
	data_disbursal.append(other_inorg_disbursal_total_count_df)
	data_disbursal.append(disbursal_count_total)
	data_disbursal.append("")
	data_disbursal.append(web_disbursal_blanks[0])
	data_disbursal.append(web_fb_disbursal_blanks[0])
	data_disbursal.append(web_google_disbursal_blanks[0])
	data_disbursal.append(web_aff_disbursal_count_sum)
	data_disbursal.append(web_vlf_disbursal_blanks[0])
	data_disbursal.append(web_2_onl_disbursal_blanks[0])
	data_disbursal.append(web_netcore_disbursal_blanks[0])
	data_disbursal.append(web_disbursal_tot)
	data_disbursal.append("")
	data_disbursal.append(market_tot_disbursal)

	print data_disbursal

	data_disbursal_amt.append(disbursal_amt_blanks[0])
	data_disbursal_amt.append(m_google_disbursal_amount)
	data_disbursal_amt.append(m_fb_disbursal_amount)
	data_disbursal_amt.append(m_inmobi_tot_disbursal_amt_count)
	data_disbursal_amt.append(m_pocket_disbursal_amt_count)
	data_disbursal_amt.append(m_LeadBolt_disbursal_amt_count)
	data_disbursal_amt.append(m_Aff_disbursal_total_amount_df)
	data_disbursal_amt.append(m_DBM_tot_disbursal_amt_count)
	data_disbursal_amt.append(m_Appnext_disbursal_amt_count)
	data_disbursal_amt.append(m_fbads_disbursal_amount)
	data_disbursal_amt.append(m_AppleSearch_disbursal_amt_count)
	data_disbursal_amt.append(madi_total_disbursal_amt)
	data_disbursal_amt.append(sk_fb_disbursal_tot_amount)
	data_disbursal_amt.append(sk_mediamath_disbursal_amt_count)
	data_disbursal_amt.append(sk_DBM_disbursal_amt_count)
	data_disbursal_amt.append(sk_google_disbursal_amount)
	data_disbursal_amt.append(sk_TikTok_disbursal_amt_count)
	data_disbursal_amt.append(sk_Affilates_disbursal_amt_total_count)
	data_disbursal_amt.append(sk_DBM_retarg_disbursal_amt_count)
	data_disbursal_amt.append(sok_total_disbursal_amt)
	data_disbursal_amt.append(seventynine_disbursal_amt_count)
	data_disbursal_amt.append(BuddyLoan_disbursal_amt_count)
	data_disbursal_amt.append(Netcore_disbursal_amt_count)
	data_disbursal_amt.append(other_inorg_disbursal_total_amount_df)
	data_disbursal_amt.append(disbursal_amount_total)
	data_disbursal_amt.append("")
	data_disbursal_amt.append(web_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_fb_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_google_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_aff_disbursal_amount_sum)
	data_disbursal_amt.append(web_vlf_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_2_onl_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_netcore_disbursal_amt_blanks[0])
	data_disbursal_amt.append(web_disbursal_amt_tot)
	data_disbursal_amt.append("")
	data_disbursal_amt.append(market_tot_disbursal_amt)

	print data_disbursal_amt


	start_col_idx = 3

	str1 = 'MIS - Dhani -'
	str2 = 	(date.today() - timedelta(days=counter)).strftime("%b")
	str3 = ' 2019.xlsx'
	filename = str1 + str2 + str3
	print (filename)

	wb = openpyxl.load_workbook('MIS - Dhani -Sept 2019.xlsx')
	sheet = wb.get_sheet_by_name('Raibow MIS')

	cntr = 0
	exclude_idx = [11,19,24,33,35]

	################# Updating App count, Disursal Count and Disbursal amount in the Rainbow file
	for i in data_app:

		if (cntr not in exclude_idx):

			sheet.cell(row=start_row_idx_1, column=start_col_idx).value = data_app[cntr]
			sheet.cell(row=start_row_idx_2, column=start_col_idx).value = data_disbursal[cntr]
			sheet.cell(row=start_row_idx_3, column=start_col_idx).value = data_disbursal_amt[cntr]
	
		start_col_idx = start_col_idx + 1
		cntr = cntr + 1


	print ("======================= SPENDS ============================")

	spends_df = pd.read_excel('Morning MIS.xlsx', sheet_name='Master Sheet')

	spends_df.reset_index(inplace=True)

	google_spends = list(spends_df.loc[spends_df["Types"] == 'Google UAC',dd_date])
	if (google_spends !=[]):
		google_spends = [x for x in google_spends if ~np.isnan(x)]
		google_spends.append(0)
	else:
		google_spends.append(0)
	
	print google_spends[0]

	fb_spends = list(spends_df.loc[spends_df["Types"] == 'Facebook',dd_date])
	if (fb_spends !=[]):
		fb_spends = [x for x in fb_spends if ~np.isnan(x)]
		fb_spends.append(0)
	else:
		fb_spends.append(0)
	
	print fb_spends[0]

	DBM_spends = list(spends_df.loc[spends_df["Types"] == 'DBM',dd_date])
	if (DBM_spends !=[]):
		DBM_spends = [x for x in DBM_spends if ~np.isnan(x)]
		DBM_spends.append(0)
	else:
		DBM_spends.append(0)

	print DBM_spends[0]

	AppNext_spends = list(spends_df.loc[spends_df["Types"] == 'AppNext',dd_date])
	if (AppNext_spends !=[]):
		AppNext_spends = [x for x in AppNext_spends if ~np.isnan(x)]
		AppNext_spends.append(0)
	else:
		AppNext_spends.append(0)

	print AppNext_spends[0]

	

	sk_fb_spends = list(spends_df.loc[spends_df["Types"] == 'Sokrati Facebook ',dd_date])
	if (sk_fb_spends !=[]):
		sk_fb_spends = [x for x in sk_fb_spends if ~np.isnan(x)]
		sk_fb_spends.append(0)
	else:
		sk_fb_spends.append(0)
	
	print sk_fb_spends[0]

	sk_MediaMath_spends = list(spends_df.loc[spends_df["Types"] == 'Sokrati MediaMath',dd_date])
	if (sk_MediaMath_spends !=[]):
		sk_MediaMath_spends = [x for x in sk_MediaMath_spends if ~np.isnan(x)]
		sk_MediaMath_spends.append(0)
	else:
		sk_MediaMath_spends.append(0)
	
	print sk_MediaMath_spends[0]


	sk_DBM_spends = list(spends_df.loc[spends_df["Types"] == 'Sokrati DBM',dd_date])
	if (sk_DBM_spends !=[]):
		sk_DBM_spends = [x for x in sk_DBM_spends if ~np.isnan(x)]
		sk_DBM_spends.append(0)
	else:
		sk_DBM_spends.append(0)
	
	print sk_DBM_spends[0]

	sk_google_spends = list(spends_df.loc[spends_df["Types"] == 'Sokrati Google UAC',dd_date])
	if (sk_google_spends !=[]):
		sk_google_spends = [x for x in sk_google_spends if ~np.isnan(x)]
		sk_google_spends.append(0)
	else:
		sk_google_spends.append(0)
	
	print sk_google_spends[0]


	sk_TikTok_spends = list(spends_df.loc[spends_df["Types"] == 'Sokrati Tiktok',dd_date])
	if (sk_TikTok_spends !=[]):
		sk_TikTok_spends = [x for x in sk_TikTok_spends if ~np.isnan(x)]
		sk_TikTok_spends.append(0)
	else:
		sk_TikTok_spends.append(0)
	
	print sk_TikTok_spends[0]

	web_google_spends = list(spends_df.loc[spends_df["Types"] == 'Google Search (Madison)',dd_date])
	if (web_google_spends !=[]):
		web_google_spends = [x for x in web_google_spends if ~np.isnan(x)]
		web_google_spends.append(0)
	else:
		web_google_spends.append(0)

	print web_google_spends[0]

	data_spends = []

	data_spends.append(google_spends[0])
	data_spends.append(fb_spends[0])
	data_spends.append(DBM_spends[0])
	data_spends.append(AppNext_spends[0])
	data_spends.append(sk_fb_spends[0])
	data_spends.append(sk_MediaMath_spends[0])
	data_spends.append(sk_DBM_spends[0])
	data_spends.append(sk_google_spends[0])
	data_spends.append(sk_TikTok_spends[0])
	data_spends.append(web_google_spends[0])

	print (data_spends)


	include_idx = [4,5,10,11,15,16,17,18,19,31]

	start_col_idx = 3
	cntr = 0

	for i in data_app:

		if (start_col_idx in include_idx):

			sheet.cell(row=start_row_idx_4, column=start_col_idx).value = data_spends[cntr]
			cntr = cntr + 1

		start_col_idx = start_col_idx + 1


	wb.save('MIS - Dhani -Sept 2019.xlsx')


	counter =  counter - 1
	index = index + 1

